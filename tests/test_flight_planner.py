"""
End-to-end and unit tests for flight_planner.py

Run with:  pytest tests/ -v
"""

import math
import subprocess
import sys
import tempfile
import textwrap
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

# Make sure the project root is importable
sys.path.insert(0, str(Path(__file__).parent.parent))
import flight_planner as fp

# ---------------------------------------------------------------------------
# Fixtures & helpers
# ---------------------------------------------------------------------------

ASSETS = Path(__file__).parent.parent / "assets"
EXAMPLE_GPX = ASSETS / "example_flight_log.gpx"
FLIGHT_DATE = datetime(2026, 4, 26)

# Minimal wind / TAS used across most tests
WIND_DIR   = 270.0
WIND_SPEED = 15.0
TAS        = 110.0
FUEL       = 6.5


def _make_gpx(waypoints: list[tuple[str, float, float, float]]) -> str:
    """Build a minimal GPX string from (name, lat, lon, ele_m) tuples."""
    pts = "\n".join(
        f'  <rtept lat="{lat}" lon="{lon}"><ele>{ele}</ele><name>{name}</name></rtept>'
        for name, lat, lon, ele in waypoints
    )
    # XML declaration must be at column 0 — no leading whitespace
    return (
        '<?xml version="1.0" encoding="utf-8"?>\n'
        '<gpx version="1.1" xmlns="http://www.topografix.com/GPX/1/1">\n'
        '  <rte>\n'
        '    <name>TEST</name>\n'
        f'{pts}\n'
        '  </rte>\n'
        '</gpx>\n'
    )


def _write_gpx(tmp_path: Path, waypoints) -> Path:
    p = tmp_path / "test_route.gpx"
    p.write_text(_make_gpx(waypoints))
    return p


def _run_cli(gpx: Path, output: Path, extra_args: list[str] | None = None) -> subprocess.CompletedProcess:
    cmd = [
        sys.executable, str(Path(__file__).parent.parent / "flight_planner.py"),
        "--gpx", str(gpx),
        "--wind-dir", str(WIND_DIR),
        "--wind-speed", str(WIND_SPEED),
        "--tas", str(TAS),
        "--fuel", str(FUEL),
        "--date", FLIGHT_DATE.strftime("%Y-%m-%d"),
        "--output", str(output),
    ] + (extra_args or [])
    return subprocess.run(cmd, capture_output=True, text=True)


# ---------------------------------------------------------------------------
# GPX parsing
# ---------------------------------------------------------------------------

class TestParseGpx:
    def test_example_gpx_waypoint_count(self):
        wpts = fp.parse_gpx(str(EXAMPLE_GPX))
        assert len(wpts) == 13

    def test_example_gpx_first_waypoint(self):
        wpts = fp.parse_gpx(str(EXAMPLE_GPX))
        assert wpts[0]["name"] == "EPMO"
        assert abs(wpts[0]["lat"] - 52.451222) < 0.001
        assert abs(wpts[0]["lon"] - 20.651861) < 0.001

    def test_example_gpx_last_waypoint(self):
        wpts = fp.parse_gpx(str(EXAMPLE_GPX))
        assert wpts[-1]["name"] == "EPMO"   # round-trip route

    def test_elevation_converted_to_meters(self, tmp_path):
        gpx = _write_gpx(tmp_path, [
            ("A", 52.0, 20.0, 500.0),
            ("B", 53.0, 20.0, 500.0),
        ])
        wpts = fp.parse_gpx(str(gpx))
        assert wpts[0]["ele_m"] == pytest.approx(500.0)

    def test_missing_name_gets_placeholder(self, tmp_path):
        content = textwrap.dedent("""\
            <?xml version="1.0"?>
            <gpx version="1.1" xmlns="http://www.topografix.com/GPX/1/1">
              <rte>
                <rtept lat="52.0" lon="20.0"><ele>100</ele></rtept>
                <rtept lat="53.0" lon="21.0"><ele>100</ele></rtept>
              </rte>
            </gpx>
        """)
        gpx = tmp_path / "noname.gpx"
        gpx.write_text(content)
        wpts = fp.parse_gpx(str(gpx))
        assert wpts[0]["name"] != ""


# ---------------------------------------------------------------------------
# Geodetic math
# ---------------------------------------------------------------------------

class TestTrueCourse:
    def test_due_north(self):
        tc = fp.true_course(50.0, 20.0, 51.0, 20.0)
        assert tc == pytest.approx(0.0, abs=0.5)

    def test_due_south(self):
        tc = fp.true_course(51.0, 20.0, 50.0, 20.0)
        assert tc == pytest.approx(180.0, abs=0.5)

    def test_due_east(self):
        tc = fp.true_course(50.0, 20.0, 50.0, 21.0)
        assert tc == pytest.approx(90.0, abs=1.0)

    def test_due_west(self):
        tc = fp.true_course(50.0, 21.0, 50.0, 20.0)
        assert tc == pytest.approx(270.0, abs=1.0)

    def test_result_always_in_0_360(self):
        # Any course should be in [0, 360)
        for lat1, lon1, lat2, lon2 in [
            (52.0, 20.0, 51.0, 19.0),
            (10.0, 10.0,  9.0, 11.0),
            (0.0,  0.0,  -1.0, -1.0),
        ]:
            tc = fp.true_course(lat1, lon1, lat2, lon2)
            assert 0.0 <= tc < 360.0


class TestDistanceNm:
    def test_short_leg(self):
        # ~1° latitude ≈ 60 NM
        d = fp.distance_nm(52.0, 20.0, 53.0, 20.0)
        assert d == pytest.approx(60.0, rel=0.02)

    def test_zero_distance(self):
        d = fp.distance_nm(52.0, 20.0, 52.0, 20.0)
        assert d == pytest.approx(0.0, abs=0.001)

    def test_symmetry(self):
        d1 = fp.distance_nm(52.0, 20.0, 53.0, 21.0)
        d2 = fp.distance_nm(53.0, 21.0, 52.0, 20.0)
        assert d1 == pytest.approx(d2, rel=1e-9)


# ---------------------------------------------------------------------------
# Wind calculations
# ---------------------------------------------------------------------------

class TestWindCalculations:
    def test_zero_wind_wca_is_zero(self):
        wca = fp.wind_correction_angle(0.0, 0.0, 90.0, 100.0)
        assert wca == pytest.approx(0.0, abs=1e-9)

    def test_zero_wind_gs_equals_tas(self):
        gs = fp.ground_speed(0.0, 0.0, 90.0, 100.0, 0.0)
        assert gs == pytest.approx(100.0, abs=1e-6)

    def test_direct_headwind_reduces_gs(self):
        # Flying east (MC=90), wind FROM east (dir=90) → direct headwind
        wca = fp.wind_correction_angle(90.0, 20.0, 90.0, 100.0)
        gs  = fp.ground_speed(90.0, 20.0, 90.0, 100.0, wca)
        assert gs < 100.0

    def test_direct_tailwind_increases_gs(self):
        # Flying east (MC=90), wind FROM west (dir=270) → direct tailwind
        wca = fp.wind_correction_angle(270.0, 20.0, 90.0, 100.0)
        gs  = fp.ground_speed(270.0, 20.0, 90.0, 100.0, wca)
        assert gs > 100.0

    def test_crosswind_wca_sign(self):
        # Wind FROM north (dir=0), flying east (MC=90)
        # Wind pushes aircraft south → need to head left (north) → WCA < 0
        # Template formula: WCA = -(ws*60*sin(wd-180-mc))/tas
        #   = -(20*60*sin(0-180-90))/100 = -(20*60*sin(-270°))/100 = -12
        wca = fp.wind_correction_angle(0.0, 20.0, 90.0, 100.0)
        assert wca < 0.0


# ---------------------------------------------------------------------------
# compute_legs
# ---------------------------------------------------------------------------

class TestComputeLegs:
    def setup_method(self):
        self.wpts = fp.parse_gpx(str(EXAMPLE_GPX))

    def test_leg_count(self):
        legs = fp.compute_legs(self.wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        assert len(legs) == 12

    def test_zero_wind_wca_all_zero(self):
        legs = fp.compute_legs(self.wpts, 0.0, 0.0, TAS, FUEL, FLIGHT_DATE)
        for leg in legs:
            assert leg["wca"] == pytest.approx(0.0, abs=0.1)

    def test_zero_wind_gs_equals_tas(self):
        legs = fp.compute_legs(self.wpts, 0.0, 0.0, TAS, FUEL, FLIGHT_DATE)
        for leg in legs:
            assert leg["gs"] == pytest.approx(TAS, abs=0.1)

    def test_mag_course_equals_tc_minus_decl(self):
        legs = fp.compute_legs(self.wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        for leg in legs:
            expected_mc = (leg["true_course"] + leg["neg_decl"]) % 360.0
            # Both values are rounded independently, so allow up to 0.2° difference
            assert leg["mag_course"] == pytest.approx(expected_mc, abs=0.2)

    def test_mag_course_in_range(self):
        legs = fp.compute_legs(self.wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        for leg in legs:
            assert 0.0 <= leg["mag_course"] < 360.0

    def test_mag_heading_in_range(self):
        legs = fp.compute_legs(self.wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        for leg in legs:
            assert 0.0 <= leg["mag_heading"] < 360.0

    def test_positive_distance(self):
        legs = fp.compute_legs(self.wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        for leg in legs:
            assert leg["dist_nm"] > 0.0

    def test_positive_time(self):
        legs = fp.compute_legs(self.wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        for leg in legs:
            assert leg["time_hr"] > 0.0

    def test_from_name_matches_waypoint(self):
        legs = fp.compute_legs(self.wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        for i, leg in enumerate(legs):
            assert leg["from_name"] == self.wpts[i]["name"]

    def test_to_name_matches_waypoint(self):
        legs = fp.compute_legs(self.wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        for i, leg in enumerate(legs):
            assert leg["to_name"] == self.wpts[i + 1]["name"]

    def test_warsaw_declination_approx(self):
        # Warsaw area 2026: expect ~6–8° East → neg_decl should be ~ -6 to -8
        legs = fp.compute_legs(self.wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        for leg in legs:
            assert -10.0 < leg["neg_decl"] < -4.0


# ---------------------------------------------------------------------------
# XLSX output (fill_plan)
# ---------------------------------------------------------------------------

class TestFillPlan:
    def _args(self, output, departure_time=""):
        class Args:
            pass
        a = Args()
        a.date = FLIGHT_DATE
        a.wind_dir = WIND_DIR
        a.wind_speed = WIND_SPEED
        a.tas = TAS
        a.fuel = FUEL
        a.aircraft_type = "C172"
        a.registration = "SP-ABC"
        a.departure_time = departure_time
        a.output = str(output)
        return a

    def test_output_file_created(self, tmp_path):
        out = tmp_path / "plan.xlsx"
        wpts = fp.parse_gpx(str(EXAMPLE_GPX))
        legs = fp.compute_legs(wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        fp.fill_plan(self._args(out), wpts, legs, str(out))
        assert out.exists()

    def test_correct_number_of_data_rows(self, tmp_path):
        out = tmp_path / "plan.xlsx"
        wpts = fp.parse_gpx(str(EXAMPLE_GPX))
        legs = fp.compute_legs(wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        fp.fill_plan(self._args(out), wpts, legs, str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        # Rows 1–3 header, row 4 blank, row 5 col headers,
        # rows 6..(6+12-1) data, last row totals
        # → data rows = total rows - 6 (header rows) - 1 (totals)
        non_empty = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
        # header(3) + col_header(1) + legs(12) + totals(1) = 17
        assert len(non_empty) == 17

    def test_header_values_present(self, tmp_path):
        out = tmp_path / "plan.xlsx"
        wpts = fp.parse_gpx(str(EXAMPLE_GPX))
        legs = fp.compute_legs(wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        fp.fill_plan(self._args(out), wpts, legs, str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        all_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert "EPMO" in all_values
        assert "C172" in all_values
        assert "SP-ABC" in all_values

    def test_totals_distance_matches_sum(self, tmp_path):
        out = tmp_path / "plan.xlsx"
        wpts = fp.parse_gpx(str(EXAMPLE_GPX))
        legs = fp.compute_legs(wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        fp.fill_plan(self._args(out), wpts, legs, str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
        totals_row = rows[-1]
        col_headers = rows[3]  # "TRASA\n(FROM)", "NKDG\nTC (°)", ...
        dist_col = next(i for i, h in enumerate(col_headers) if h and "Dist" in str(h))
        total_dist = totals_row[dist_col]
        expected = round(sum(l["dist_nm"] for l in legs), 1)
        assert total_dist == pytest.approx(expected, abs=0.1)

    def test_totals_fuel_matches_sum(self, tmp_path):
        out = tmp_path / "plan.xlsx"
        wpts = fp.parse_gpx(str(EXAMPLE_GPX))
        legs = fp.compute_legs(wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        fp.fill_plan(self._args(out), wpts, legs, str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
        totals_row = rows[-1]
        col_headers = rows[3]
        fuel_col = next(i for i, h in enumerate(col_headers) if h and "FUEL" in str(h))
        assert totals_row[fuel_col] == pytest.approx(round(sum(l["fuel"] for l in legs), 1), abs=0.1)

    def test_waypoint_names_in_data_rows(self, tmp_path):
        out = tmp_path / "plan.xlsx"
        wpts = fp.parse_gpx(str(EXAMPLE_GPX))
        legs = fp.compute_legs(wpts, WIND_DIR, WIND_SPEED, TAS, FUEL, FLIGHT_DATE)
        fp.fill_plan(self._args(out), wpts, legs, str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
        data_rows = rows[4:-1]   # skip headers + totals
        names_in_sheet = {r[0] for r in data_rows}
        expected_from_names = {l["from_name"] for l in legs}
        assert expected_from_names == names_in_sheet


# ---------------------------------------------------------------------------
# CLI integration
# ---------------------------------------------------------------------------

class TestCli:
    def test_cli_runs_successfully(self, tmp_path):
        out = tmp_path / "plan.xlsx"
        result = _run_cli(EXAMPLE_GPX, out)
        assert result.returncode == 0
        assert out.exists()

    def test_cli_output_contains_summary(self, tmp_path):
        out = tmp_path / "plan.xlsx"
        result = _run_cli(EXAMPLE_GPX, out)
        assert "Total distance" in result.stdout
        assert "Total time" in result.stdout
        assert "Total fuel" in result.stdout

    def test_cli_optional_args(self, tmp_path):
        out = tmp_path / "plan.xlsx"
        result = _run_cli(EXAMPLE_GPX, out, [
            "--aircraft-type", "PA28",
            "--registration", "SP-XYZ",
            "--departure-time", "10:30",
        ])
        assert result.returncode == 0

    def test_cli_rejects_zero_tas(self, tmp_path):
        out = tmp_path / "plan.xlsx"
        result = _run_cli(EXAMPLE_GPX, out, ["--tas", "0"])
        # argparse overrides --tas; check it fails
        # We re-run with subprocess directly to override --tas
        cmd = [
            sys.executable, str(Path(__file__).parent.parent / "flight_planner.py"),
            "--gpx", str(EXAMPLE_GPX),
            "--wind-dir", "270", "--wind-speed", "15",
            "--tas", "0", "--fuel", "6.5",
            "--date", "2026-04-26",
            "--output", str(out),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True)
        assert result.returncode != 0

    def test_cli_rejects_negative_wind_speed(self, tmp_path):
        cmd = [
            sys.executable, str(Path(__file__).parent.parent / "flight_planner.py"),
            "--gpx", str(EXAMPLE_GPX),
            "--wind-dir", "270", "--wind-speed", "-5",
            "--tas", "110", "--fuel", "6.5",
            "--date", "2026-04-26",
            "--output", str(tmp_path / "plan.xlsx"),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True)
        assert result.returncode != 0

    def test_cli_rejects_wind_dir_out_of_range(self, tmp_path):
        cmd = [
            sys.executable, str(Path(__file__).parent.parent / "flight_planner.py"),
            "--gpx", str(EXAMPLE_GPX),
            "--wind-dir", "400", "--wind-speed", "10",
            "--tas", "110", "--fuel", "6.5",
            "--date", "2026-04-26",
            "--output", str(tmp_path / "plan.xlsx"),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True)
        assert result.returncode != 0

    def test_cli_single_leg_route(self, tmp_path):
        gpx = _write_gpx(tmp_path, [
            ("EPMO", 52.451222, 20.651861, 105.0),
            ("DEST", 53.000000, 21.000000, 300.0),
        ])
        out = tmp_path / "plan.xlsx"
        result = _run_cli(gpx, out)
        assert result.returncode == 0
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
        # header(3) + col_header(1) + 1 leg + totals(1) = 6
        assert len(rows) == 6
