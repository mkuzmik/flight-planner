#!/usr/bin/env python3
"""
flight_planner.py — Generate an operational flight plan XLSX from a GPX route.

Usage:
    python flight_planner.py --gpx route.gpx --wind-dir 270 --wind-speed 15 \
        --tas 110 --fuel 6.5 [options]
"""

import argparse
import math
import sys
from datetime import datetime, date as date_type
from pathlib import Path

import gpxpy
import numpy as np
import ppigrf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from geographiclib.geodesic import Geodesic

GEOD = Geodesic.WGS84


# ---------------------------------------------------------------------------
# GPX parsing
# ---------------------------------------------------------------------------

def parse_gpx(gpx_path: str) -> list[dict]:
    """Return list of waypoints from a GPX route or track."""
    with open(gpx_path, encoding="utf-8") as f:
        gpx = gpxpy.parse(f)

    waypoints = []

    for route in gpx.routes:
        for pt in route.points:
            waypoints.append({
                "name": (pt.name or f"WPT{len(waypoints) + 1}").strip(),
                "lat": pt.latitude,
                "lon": pt.longitude,
                "ele_m": pt.elevation or 0.0,
            })

    if not waypoints:
        for track in gpx.tracks:
            for seg in track.segments:
                for pt in seg.points:
                    waypoints.append({
                        "name": (pt.name or f"WPT{len(waypoints) + 1}").strip(),
                        "lat": pt.latitude,
                        "lon": pt.longitude,
                        "ele_m": pt.elevation or 0.0,
                    })

    return waypoints


# ---------------------------------------------------------------------------
# Navigation calculations
# ---------------------------------------------------------------------------

def true_course(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Forward azimuth (true course) in degrees [0, 360)."""
    result = GEOD.Inverse(lat1, lon1, lat2, lon2)
    return result["azi1"] % 360.0


def distance_nm(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Great-circle distance in nautical miles."""
    result = GEOD.Inverse(lat1, lon1, lat2, lon2)
    return result["s12"] / 1852.0


def magnetic_declination(lat: float, lon: float, alt_m: float, dt: datetime) -> float:
    """Magnetic declination in degrees (East-positive) using IGRF-14."""
    alt_km = alt_m / 1000.0
    Be, Bn, _ = ppigrf.igrf(lon, lat, alt_km, dt)
    return math.degrees(math.atan2(float(np.asarray(Be).flat[0]),
                                   float(np.asarray(Bn).flat[0])))


def wind_correction_angle(wind_dir: float, wind_speed: float,
                          mag_course: float, tas: float) -> float:
    """Wind correction angle in degrees (same approximation as the template).

    The template uses:
        WCA = -(wind_speed * 60 * sin(wind_dir - 180 - mag_course)) / TAS

    where 60 ≈ 180/π is a mental-math approximation of the arcsin linearisation.
    """
    angle_rad = math.radians(wind_dir - 180.0 - mag_course)
    return -(wind_speed * 60.0 * math.sin(angle_rad)) / tas


def ground_speed(wind_dir: float, wind_speed: float,
                 mag_course: float, tas: float, wca_deg: float) -> float:
    """Ground speed in knots (same formula as the template)."""
    angle_rad = math.radians(wind_dir - 180.0 - mag_course)
    return tas * math.cos(math.radians(wca_deg)) + wind_speed * math.cos(angle_rad)


# ---------------------------------------------------------------------------
# Leg computation
# ---------------------------------------------------------------------------

def compute_legs(waypoints: list[dict], wind_dir: float, wind_speed: float,
                 tas: float, fuel_gph: float, dt: datetime) -> list[dict]:
    """Compute navigation data for every leg of the route."""
    legs = []
    for i in range(1, len(waypoints)):
        p1 = waypoints[i - 1]
        p2 = waypoints[i]

        mid_lat = (p1["lat"] + p2["lat"]) / 2.0
        mid_lon = (p1["lon"] + p2["lon"]) / 2.0
        mid_alt_m = (p1["ele_m"] + p2["ele_m"]) / 2.0

        tc = true_course(p1["lat"], p1["lon"], p2["lat"], p2["lon"])
        dist = distance_nm(p1["lat"], p1["lon"], p2["lat"], p2["lon"])
        decl = magnetic_declination(mid_lat, mid_lon, mid_alt_m, dt)

        neg_decl = -decl                          # value for template col C
        mc = (tc + neg_decl) % 360.0             # magnetic course

        wca = wind_correction_angle(wind_dir, wind_speed, mc, tas)
        mh = (mc + wca) % 360.0                  # magnetic heading

        gs = max(ground_speed(wind_dir, wind_speed, mc, tas, wca), 1.0)

        time_hr = dist / gs
        fuel = time_hr * fuel_gph

        legs.append({
            "from_name": p1["name"],
            "to_name": p2["name"],
            "true_course": round(tc, 1),
            "neg_decl": round(neg_decl, 1),
            "mag_course": round(mc, 1),
            "wca": round(wca, 1),
            "mag_heading": round(mh % 360.0, 1),
            "dist_nm": round(dist, 1),
            "gs": round(gs, 1),
            "time_hr": time_hr,
            "time_excel": time_hr / 24.0,   # Excel serial fraction
            "fuel": round(fuel, 1),
        })
    return legs


# ---------------------------------------------------------------------------
# Excel output — clean sheet, no template dependency
# ---------------------------------------------------------------------------

# Colour palette
_DARK_BLUE   = "1F3864"
_MID_BLUE    = "2E75B6"
_LIGHT_BLUE  = "BDD7EE"
_LIGHT_GREY  = "F2F2F2"
_WHITE       = "FFFFFF"

def _hdr_font(white: bool = True) -> Font:
    return Font(name="Calibri", bold=True, color=_WHITE if white else "000000", size=10)

def _body_font() -> Font:
    return Font(name="Calibri", size=10)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _thin_border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _centre() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def fill_plan(args, waypoints: list[dict], legs: list[dict], output_path: str) -> None:
    """Write a clean flight-plan XLSX (no template required)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flight Plan"

    # ── helpers ────────────────────────────────────────────────────────────
    def hdr_cell(row, col, value, bg=_DARK_BLUE, white_text=True):
        c = ws.cell(row, col, value)
        c.font = _hdr_font(white_text)
        c.fill = _fill(bg)
        c.alignment = _centre()
        c.border = _thin_border()
        return c

    def label_cell(row, col, value):
        c = ws.cell(row, col, value)
        c.font = Font(name="Calibri", bold=True, size=10)
        c.alignment = _left()
        return c

    def value_cell(row, col, value, fmt=None, bg=None, centre=False):
        c = ws.cell(row, col, value)
        c.font = _body_font()
        c.border = _thin_border()
        c.alignment = _centre() if centre else _left()
        if fmt:
            c.number_format = fmt
        if bg:
            c.fill = _fill(bg)
        return c

    # ── Section 1: flight header ────────────────────────────────────────────
    row = 1
    ws.merge_cells(f"A{row}:J{row}")
    c = ws.cell(row, 1, "OPERACYJNY PLAN LOTU / OPERATIONAL FLIGHT PLAN")
    c.font = Font(name="Calibri", bold=True, size=13, color=_WHITE)
    c.fill = _fill(_DARK_BLUE)
    c.alignment = _centre()

    def _kv_pair(row, col, label, value):
        """Write a label cell and adjacent value cell (no merges)."""
        lc = ws.cell(row, col, label)
        lc.font = Font(name="Calibri", bold=True, size=9, color=_WHITE)
        lc.fill = _fill(_MID_BLUE)
        lc.alignment = _centre()
        vc = ws.cell(row, col + 1, value)
        vc.font = _body_font()
        vc.alignment = _centre()
        vc.border = _thin_border()

    row = 2
    _kv_pair(row, 1, "DATA / DATE",          args.date.strftime("%Y-%m-%d"))
    _kv_pair(row, 3, "ODLOT / DEP",          waypoints[0]["name"])
    _kv_pair(row, 5, "DOLOT / DEST",         waypoints[-1]["name"])
    _kv_pair(row, 7, "TYP / TYPE",           args.aircraft_type or "—")
    _kv_pair(row, 9, "ZNAKI / REG",          args.registration or "—")

    row = 3
    _kv_pair(row, 1, "WIATR DIR (°)",        args.wind_dir)
    _kv_pair(row, 3, "WIATR SPD (kt)",       args.wind_speed)
    _kv_pair(row, 5, "TAS (kt)",             args.tas)
    _kv_pair(row, 7, "FUEL (gal/h)",         args.fuel)
    _kv_pair(row, 9, "GODZ. START",          args.departure_time or "—")

    # ── Section 2: route table ───────────────────────────────────────────────
    row = 5
    COLS = [
        ("TRASA\n(FROM)",   16, "left"),
        ("NKDG\nTC (°)",     7, "center"),
        ("(-) DM\nDecl (°)", 7, "center"),
        ("NKDM\nMC (°)",     7, "center"),
        ("(-) KZ\nWCA (°)",  7, "center"),
        ("KM\nMH (°)",       7, "center"),
        ("S\nDist (NM)",     8, "center"),
        ("W\nGS (kt)",       7, "center"),
        ("t\nTime",          8, "center"),
        ("FUEL\n(gal)",      7, "center"),
    ]

    # Header row
    ws.row_dimensions[row].height = 28
    for col_idx, (label, width, align) in enumerate(COLS, start=1):
        c = hdr_cell(row, col_idx, label)
        c.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for i, leg in enumerate(legs):
        row += 1
        bg = _LIGHT_GREY if i % 2 == 0 else _WHITE
        ws.row_dimensions[row].height = 14

        def dv(col_idx, val, fmt=None):
            value_cell(row, col_idx, val, fmt=fmt, bg=bg, centre=(col_idx > 1))

        dv(1,  leg["from_name"])
        dv(2,  leg["true_course"],  "0.0")
        dv(3,  leg["neg_decl"],     "+0.0;-0.0;0.0")
        dv(4,  leg["mag_course"],   "0.0")
        dv(5,  leg["wca"],          "+0.0;-0.0;0.0")
        dv(6,  leg["mag_heading"],  "0.0")
        dv(7,  leg["dist_nm"],      "0.0")
        dv(8,  leg["gs"],           "0.0")
        dv(9,  leg["time_excel"],   "[m]:ss")
        dv(10, leg["fuel"],         "0.0")

    # Totals row
    row += 1
    total_time_hr = sum(l["time_hr"] for l in legs)
    total_fuel    = sum(l["fuel"] for l in legs)
    total_dist    = sum(l["dist_nm"] for l in legs)

    hdr_cell(row, 1, "TOTAL", bg=_MID_BLUE)
    for col_idx in range(2, 11):
        ws.cell(row, col_idx).fill = _fill(_LIGHT_BLUE)
        ws.cell(row, col_idx).border = _thin_border()
        ws.cell(row, col_idx).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row, col_idx).alignment = _centre()

    ws.cell(row, 7).value = round(total_dist, 1)
    ws.cell(row, 7).number_format = "0.0"
    ws.cell(row, 9).value = total_time_hr / 24.0
    ws.cell(row, 9).number_format = "[h]:mm"
    ws.cell(row, 10).value = round(total_fuel, 1)
    ws.cell(row, 10).number_format = "0.0"

    # Freeze header rows + header column
    ws.freeze_panes = "B6"

    wb.save(output_path)

    # ── Console summary ──────────────────────────────────────────────────────
    hours   = int(total_time_hr)
    minutes = round((total_time_hr - hours) * 60)
    n_legs  = len(legs)
    print(f"✓ Flight plan saved: {output_path}")
    print(f"  Route:          {waypoints[0]['name']} → {waypoints[-1]['name']}"
          f"  ({n_legs} legs)")
    print(f"  Total distance: {total_dist:.1f} NM")
    print(f"  Total time:     {hours}h {minutes:02d}m")
    print(f"  Total fuel:     {total_fuel:.1f} gal")
    print(f"  Wind:           {args.wind_dir:.0f}° / {args.wind_speed:.0f} kt  "
          f"TAS: {args.tas:.0f} kt")
    print(f"  Mag. decl.:     {legs[0]['neg_decl'] * -1:+.1f}° (first leg midpoint)")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate an operational flight plan XLSX from a GPX route file.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--gpx", required=True,
                        help="Path to GPX file containing the route")
    parser.add_argument("--wind-dir", type=float, required=True, metavar="DEG",
                        help="Wind direction in degrees (FROM)")
    parser.add_argument("--wind-speed", type=float, required=True, metavar="KT",
                        help="Wind speed in knots")
    parser.add_argument("--tas", type=float, required=True, metavar="KT",
                        help="True airspeed in knots")
    parser.add_argument("--fuel", type=float, required=True, metavar="GAL/H",
                        help="Fuel consumption in US gallons per hour")
    parser.add_argument("--date", metavar="YYYY-MM-DD",
                        default=date_type.today().strftime("%Y-%m-%d"),
                        help="Flight date (used for magnetic declination)")
    parser.add_argument("--aircraft-type", default="", metavar="TYPE",
                        help="Aircraft type (e.g. C172)")
    parser.add_argument("--registration", default="", metavar="REG",
                        help="Aircraft registration (e.g. SP-ABC)")
    parser.add_argument("--departure-time", default="", metavar="HH:MM",
                        help="Planned departure time in UTC")
    parser.add_argument("--output", default="flight_plan.xlsx", metavar="FILE",
                        help="Output XLSX file path")

    args = parser.parse_args()

    # Validate
    errors = []
    if not (0 <= args.wind_dir <= 360):
        errors.append("--wind-dir must be between 0 and 360")
    if args.wind_speed < 0:
        errors.append("--wind-speed must be ≥ 0")
    if args.tas <= 0:
        errors.append("--tas must be > 0")
    if args.fuel < 0:
        errors.append("--fuel must be ≥ 0")
    if errors:
        parser.error("\n".join(errors))

    # Parse date → datetime for ppigrf
    try:
        args.date = datetime.strptime(args.date, "%Y-%m-%d")
    except ValueError:
        parser.error("--date must be in YYYY-MM-DD format")

    waypoints = parse_gpx(args.gpx)
    if len(waypoints) < 2:
        parser.error("GPX file must contain at least 2 waypoints")

    print(f"Parsed {len(waypoints)} waypoints ({len(waypoints) - 1} legs)")

    legs = compute_legs(
        waypoints, args.wind_dir, args.wind_speed,
        args.tas, args.fuel, args.date,
    )
    fill_plan(args, waypoints, legs, args.output)


if __name__ == "__main__":
    main()
